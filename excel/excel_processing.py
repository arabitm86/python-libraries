import openpyxl as oxl
from openpyxl.chart import BarChart, Reference

def process_workbook (file_in, file_out, sheet_name, chart_location):
    
    workbook = oxl.load_workbook(file_in)
    sheet = workbook[sheet_name]

    for row in range (2, sheet.max_row + 1):
        print (sheet.cell(row,3).value)
        corrected_price = sheet.cell(row,3).value * 0.9
        
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price
        print (corrected_price_cell.value)
        
    #selects all the values in all the rows
    #Reference (sheet, min_row=2, max_row=sheet.max_row)
        
    values = Reference (sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, chart_location)
    workbook.save(file_out)
     
    
    
    